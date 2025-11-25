import copy
import warnings

from openpyxl import load_workbook
from pynput.keyboard import Controller

from BTD6_Key_and_Mouse_Controls import move2, left_click, left_down, left_up, right_down, right_up, scroll_up, \
    scroll_down, delay, key_press
from BTD6_Level_Control import wait_level, change_game_speed, GameRestartException, toggle_autostart, \
    LevelRetryException, resovle_keybind
from BTD6_Monkeys_Controls import monkey_place, monkey_place_desperate, monkey_upgrade, monkey_sell, \
    monkey_upgrade_desperate, monkey_retarget, monkey_retarget_aim, monkey_overclock, use_geraldo_tool, \
    activate_map_item
from BTD6_TraceLogs import write_game_log
from config import config, global_text

keyboard = Controller()


# 定义猴子状态的存储结构
class Monkey:
    def __init__(self, monkey_id, x=0, y=0, path1_level=0, path2_level=0, path3_level=0, key='None'):
        self.monkey_id = monkey_id
        self.x = x
        self.y = y
        self.path1_level = path1_level
        self.path2_level = path2_level
        self.path3_level = path3_level
        self.key = key

    def __str__(self):
        return (f"Monkey(id={self.monkey_id}, x={self.x}, y={self.y}, "
                f"path1_level={self.path1_level}, path2_level={self.path2_level}, "
                f"path3_level={self.path3_level}, key={self.key})")

    def __repr__(self):
        return (f"Monkey(id={self.monkey_id}, x={self.x}, y={self.y}, "
                f"path1_level={self.path1_level}, path2_level={self.path2_level}, "
                f"path3_level={self.path3_level}, key={self.key})")

    def upgrade(self, target_path1, target_path2, target_path3):
        """
        升级猴子的路径
        """
        upgrade_path1 = max(0, target_path1 - self.path1_level)
        upgrade_path2 = max(0, target_path2 - self.path2_level)
        upgrade_path3 = max(0, target_path3 - self.path3_level)

        # 更新等级
        self.path1_level += upgrade_path1
        self.path2_level += upgrade_path2
        self.path3_level += upgrade_path3

        # print(f"Upgrading {self.monkey_id}: Path1 +{upgrade_path1}, Path2 +{upgrade_path2}, Path3 +{upgrade_path3}")
        # print(f"New levels for {self.monkey_id}: ({self.path1_level}, {self.path2_level}, {self.path3_level})")

    def place(self, x, y):
        """
        放置猴子到指定位置
        """
        self.x = x
        self.y = y
        # print(f"Placed {self.monkey_id} at ({self.x}, {self.y})")

    def write(self, x, y, path1_level, path2_level, path3_level):
        """
        在猴子因位移或变身时手动修改数据
        """
        self.x = x
        self.y = y
        self.path1_level = path1_level
        self.path2_level = path2_level
        self.path3_level = path3_level


# 定义一个猴子状态的全局记录
monkeys = {}
monkeys_temp = {}


# 从 xlsx 文件读取初始状态
def load_initial_state(file_path):
    if config.SCRIPT_STOP == 1:
        return

    seen_ids = set()
    try:
        # 屏蔽openpyxl格式警告
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            wb = load_workbook(file_path, read_only=True, data_only=True)

        # 明确指定第一个工作表
        worksheet = wb.worksheets[0] if wb.worksheets else None
        if not worksheet:
            raise ValueError("Excel文件没有有效的工作表")

        # 安全获取标题行
        try:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
        except StopIteration:
            raise ValueError("工作表为空")

        # 过滤空标题
        field_names = [cell.value.strip() if cell.value else f"未知列_{i}"
                       for i, cell in enumerate(header_row, 1)]

        # 读取所有行数据到列表中
        all_rows = []
        for row in worksheet.iter_rows(min_row=2):
            row_values = [cell.value if cell.value is not None else "" for cell in row]
            row_dict = {field_names[i]: str(val).strip() for i, val in enumerate(row_values) if i < len(field_names)}
            all_rows.append(row_dict)

        # 现在可以关闭工作簿了
        # wb.close()

        initial_param = None
        for row in worksheet.iter_rows(min_row=2):
            # 处理可能为None的单元格值
            row_values = [cell.value if cell.value is not None else "" for cell in row]
            row_dict = {field_names[i]: str(val).strip() for i, val in enumerate(row_values) if i < len(field_names)}

            if row_dict.get("指令") == "地图难度":
                initial_param = row_dict.get("首要参数")
                break

        # 检查是否是数字格式的难度设置，例如"1,80"
        if initial_param and "," in initial_param:
            try:
                # 分割字符串并转换为整数
                start_level, max_level = map(int, initial_param.split(","))
                config.START_GAME_LEVEL = start_level
                config.MAX_GAME_LEVEL = max_level
                config.NOW_DIFFICULTY = "自定义"

            except ValueError:
                raise ValueError(f"请输入正确的地图难度或数字格式(如'1,80')。当前输入：{initial_param}")
        else:

            if initial_param in ("简单", "仅初级"):
                config.START_GAME_LEVEL, config.MAX_GAME_LEVEL = 1, 40
            elif initial_param in ("简单快速路径", "仅初级快速路径"):
                config.START_GAME_LEVEL, config.MAX_GAME_LEVEL = 10, 40
            elif initial_param in ("放气", "（占位符1）"):
                config.START_GAME_LEVEL, config.MAX_GAME_LEVEL = 30, 30
            elif initial_param in ("中级", "仅限军事", "天启", "相反"):
                config.START_GAME_LEVEL, config.MAX_GAME_LEVEL = 1, 60
            elif initial_param in ("中级快速路径", "仅限军事快速路径", "天启快速路径", "相反快速路径"):
                config.START_GAME_LEVEL, config.MAX_GAME_LEVEL = 15, 60
            elif initial_param in ("困难", "只限魔法猴", "双倍生命值MOAB", "现金减半", "替代气球回合"):
                config.START_GAME_LEVEL, config.MAX_GAME_LEVEL = 3, 80
            elif initial_param in ("困难快速路径", "只限魔法猴快速路径", "双倍生命值MOAB快速路径", "现金减半快速路径", "替代气球回合快速路径"):
                config.START_GAME_LEVEL, config.MAX_GAME_LEVEL = 22, 80
            elif initial_param in ("极难", "点击"):
                config.START_GAME_LEVEL, config.MAX_GAME_LEVEL = 6, 100
            elif initial_param in ("极难快速路径", "（占位符2）"):
                config.START_GAME_LEVEL, config.MAX_GAME_LEVEL = 29, 100
            elif initial_param in "测试":
                config.START_GAME_LEVEL, config.MAX_GAME_LEVEL = 99, 100
            else:
                raise ValueError(f"请输入正确的地图难度。当前输入：{initial_param}")

            config.NOW_DIFFICULTY = initial_param

        config.NOW_GAME_LEVEL = config.START_GAME_LEVEL

        print(f'当前地图难度：{initial_param}，最小关卡{config.START_GAME_LEVEL}，最大关卡{config.MAX_GAME_LEVEL}')

        if config.LOG_FILE_GRANULARITY >= 2:
            write_game_log(
                f"目前地图难度：{initial_param}",
                config.CUSTOM_SAVE_PATH
            )

        for row in worksheet.iter_rows(min_row=2):
            row_values = [cell.value if cell.value is not None else "" for cell in row]
            row_dict = {field_names[i]: str(val).strip() for i, val in enumerate(row_values) if i < len(field_names)}

            monkey_id = row_dict.get("id")
            if monkey_id and monkey_id not in seen_ids:
                x = int(row_dict["x"]) if row_dict["x"].strip() else 0
                y = int(row_dict["y"]) if row_dict["y"].strip() else 0
                path1_level = 0
                path2_level = 0
                path3_level = 0
                key = row_dict.get("按键", "")
                monkeys[monkey_id] = Monkey(monkey_id, x, y, path1_level, path2_level, path3_level, key)
                seen_ids.add(monkey_id)

        # 现在可以关闭工作簿了
        wb.close()

    except Exception as e:
        print(f"加载初始状态时出错: {e}")
        if 'wb' in locals():
            wb.close()
        raise


# 执行单条指令
def execute_command(row):
    """
    执行单条指令的逻辑
    """

    if config.SCRIPT_STOP == 1:
        return

    command = row["指令"]
    initial_param = row["首要参数"]
    monkey_id = row.get("id")
    x = row.get("x")
    y = row.get("y")
    target_path1 = row.get("上路")
    target_path2 = row.get("中路")
    target_path3 = row.get("下路")
    key = row.get("按键")

    # 如果initial_param是整数字符串，如"-1","1564"，则转化为int。如果为字符串但是并非纯数字，则保持为str
    if isinstance(initial_param, str):  # 如果是字符串
        if initial_param.strip() == "":  # 如果是空字符串
            initial_param = 0
        elif initial_param.lstrip('-').isdigit():  # 如果是数字字符串（支持负数）
            initial_param = int(initial_param)  # 转换为整数
        # 否则保持为原始字符串

    elif initial_param is None:  # 如果是 None
        initial_param = 0

    x = int(x) if x else 0
    y = int(y) if y else 0
    target_path1 = int(target_path1) if target_path1 else 0
    target_path2 = int(target_path2) if target_path2 else 0
    target_path3 = int(target_path3) if target_path3 else 0

    if command == "放置猴子":
        print(f"执行指令：放置猴子 (关卡: {config.NOW_GAME_LEVEL})")
        if not monkey_id or not x or not y:
            raise ValueError(f"放置猴子时缺少必要参数：id={monkey_id}, x={x}, y={y}")

        monkeys[monkey_id].place(x, y)
        monkey_place(monkeys[monkey_id], target_path1, target_path2, target_path3)

        if target_path1 or target_path2 or target_path3:  # 如果需要升级，则列表也需要更新
            monkeys[monkey_id].upgrade(target_path1, target_path2, target_path3)

    elif command == "升级猴子":
        print(f"执行指令：升级猴子 (关卡: {config.NOW_GAME_LEVEL})")
        if not monkey_id:
            raise ValueError(f"升级猴子时缺少必要参数：id={monkey_id}")

        monkey_upgrade(
            monkeys[monkey_id],
            target_path1 - monkeys[monkey_id].path1_level,
            target_path2 - monkeys[monkey_id].path2_level,
            target_path3 - monkeys[monkey_id].path3_level
        )
        monkeys[monkey_id].upgrade(target_path1, target_path2, target_path3)

    elif command == "急切放置猴子":
        print(f"执行指令：急切放置猴子 (关卡: {config.NOW_GAME_LEVEL})")
        if not monkey_id or not x or not y:
            raise ValueError(f"放置猴子时缺少必要参数：id={monkey_id}, x={x}, y={y}")

        monkey_place_desperate(monkeys[monkey_id], initial_param)
        monkeys[monkey_id].place(x, y)

    elif command == "急切升级猴子":
        print(f"执行指令：急切升级猴子 (关卡: {config.NOW_GAME_LEVEL})")
        if not monkey_id:
            raise ValueError(f"升级猴子时缺少必要参数：id={monkey_id}")

        monkey_upgrade_desperate(
            monkeys[monkey_id],
            target_path1 - monkeys[monkey_id].path1_level,
            target_path2 - monkeys[monkey_id].path2_level,
            target_path3 - monkeys[monkey_id].path3_level,
            initial_param
        )
        monkeys[monkey_id].upgrade(target_path1, target_path2, target_path3)

    elif command == "出售猴子":
        print(f"执行指令：出售猴子 (关卡: {config.NOW_GAME_LEVEL})")
        if not monkey_id:
            raise ValueError(f"出售猴子时缺少必要参数：id={monkey_id}")
        monkey_sell(monkeys[monkey_id])

    elif command == "改变目标":
        print(f"执行指令：改变猴子目标 (关卡: {config.NOW_GAME_LEVEL})")
        if not monkey_id:
            raise ValueError(f"改变猴子目标时缺少必要参数：id={monkey_id}, 首要参数={initial_param}")
        monkey_retarget(monkeys[monkey_id], initial_param)

    elif command == "改变瞄准":
        print(f"执行指令：改变猴子瞄准 (关卡: {config.NOW_GAME_LEVEL})")
        if not monkey_id:
            raise ValueError(f"改变猴子瞄准时缺少必要参数：id={monkey_id}, x={x}, y={y}")
        monkey_retarget_aim(monkeys[monkey_id], initial_param, x, y)

    # elif command == "使用道具":

    # 由于有时多个道具一起使用，因此在外层处理。此处仅在逻辑上占位
    # use_geraldo_tool(monkeys[monkey_id])

    elif command == "超频猴子":
        print(f"执行指令：超频猴子 (关卡: {config.NOW_GAME_LEVEL})")
        if not monkey_id or not key:
            raise ValueError(f"超频猴子时缺少必要参数：id={monkey_id}, key={key}")
        monkey_overclock(monkeys[monkey_id], key)

    elif command == "切换自动开始":
        print(f"执行指令：切换自动开始 (关卡: {config.NOW_GAME_LEVEL})")
        if initial_param != 0 and initial_param != 1:
            raise ValueError(f"切换自动开始时缺少必要参数：id={monkey_id}, 首要参数={initial_param}")
        toggle_autostart(initial_param)

    elif command == "关前停顿":
        print(f"执行指令：关前停顿 (关卡: {config.NOW_GAME_LEVEL})")
        config.STOP_HERE += 1

    elif command == "地图互动":
        print(f"执行指令：地图元素互动 (关卡: {config.NOW_GAME_LEVEL})")
        if not x and not y and not target_path1 and not target_path2:
            raise ValueError(f"地图元素互动时缺少必要参数：x1={x}, y1={y}, x2={target_path1}, y2={target_path2}")
        activate_map_item(x, y, target_path1, target_path2)

    elif command == "延时":
        print(f"执行指令：延时 {initial_param}毫秒")
        if not initial_param:
            raise ValueError(f"延时缺少必要参数：首要参数={initial_param}")
        delay(initial_param)

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：延时{initial_param}毫秒",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "键盘点击":
        if initial_param == 0:
            initial_param = 1
        print(f"执行指令：键盘点击{key}键")
        if not key:
            raise ValueError(f"键盘点击缺少必要参数：key={key}")
        key_press(key, initial_param)

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：按下{key}键{initial_param}次",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "鼠标移动":
        print(f"执行指令：鼠标移动至{x},{y}")
        if not x or not y:
            raise ValueError(f"鼠标移动缺少必要参数：x={x}, y={y}")
        move2(x, y)

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：移动鼠标至{x}, {y}",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "鼠标点击":
        if initial_param == 0:  # 若不指明次数，则设定默认次数
            initial_param = 1
        print(f"执行指令：鼠标点击{initial_param}次（如无数据默认为1）")
        left_click(initial_param)

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：左键单击{initial_param}次",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "鼠标左键按下":
        print(f"执行指令：鼠标左键按下")
        left_down()

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：鼠标左键按下",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "鼠标左键弹起":
        print(f"执行指令：鼠标左键弹起")
        left_up()

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：鼠标左键弹起",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "鼠标右键按下":
        print(f"执行指令：鼠标右键按下")
        right_down()

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：鼠标右键按下",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "鼠标右键弹起":
        print(f"执行指令：鼠标右键弹起")
        right_up()

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：鼠标右键弹起",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "鼠标左键按下拖动弹起":
        if not x or not y or not target_path1 or not target_path2:
            raise ValueError(f"鼠标左键按下拖动弹起时缺少必要参数：x1={x}, y1={y}, x2={target_path1}, y2={target_path2}")

        print(f"执行指令：鼠标移动至{x},{y}并按下左键，延时{initial_param}毫秒后，拖动至{target_path1},{target_path2}，然后弹起")
        move2(x, y)
        left_down()
        if initial_param:
            delay(initial_param)
        move2(target_path1, target_path2)
        left_up()

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：鼠标左键从{x},{y}按下，延时{initial_param}毫秒后，拖动至{target_path1},{target_path2}后弹起",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "鼠标右键按下拖动弹起":
        if not x or not y or not target_path1 or not target_path2:
            raise ValueError(f"鼠标右键按下拖动弹起时缺少必要参数：x1={x}, y1={y}, x2={target_path1}, y2={target_path2}")

        print(f"执行指令：鼠标移动至{x},{y}并按下右键，延时{initial_param}毫秒后，拖动至{target_path1},{target_path2}，然后弹起")
        move2(x, y)
        right_down()
        if initial_param:
            delay(initial_param)
        move2(target_path1, target_path2)
        right_up()

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：鼠标右键从{x},{y}按下，延时{initial_param}毫秒后，拖动至{target_path1},{target_path2}后弹起",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "鼠标上滚":
        if initial_param == 0:  # 若不指明次数，则设定默认次数
            initial_param = 1
        print(f"执行指令：鼠标上滚{initial_param}次（如无数据默认为1）")
        scroll_up(initial_param)

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：鼠标上滚{initial_param}次",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "鼠标下滚":
        if initial_param == 0:  # 若不指明次数，则设定默认次数
            initial_param = 1
        print(f"执行指令：鼠标下滚{initial_param}次（如无数据默认为1）")
        scroll_down(initial_param)

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：鼠标下滚{initial_param}次",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "延时移动点击":
        move2(x, y)

        if initial_param:
            delay(initial_param)
        print(f"执行指令：延时{initial_param}毫秒后，鼠标移动至{x},{y}，点击1次")
        left_click()

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：延时{initial_param}毫秒后，鼠标移动至{x},{y}，左键单击1次",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "延时按键":
        if initial_param:
            delay(initial_param)
        print(f"执行指令：延时{initial_param}毫秒后，按键{key}1次")
        key_press(key, 1)

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"自定义动作：延时{initial_param}毫秒后，鼠标移动至{x},{y}，左键单击1次",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "猴子状态被强制改变":
        print(f"执行指令：手动修改猴子状态")
        monkeys[monkey_id].write(x, y, target_path1, target_path2, target_path3)

        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"猴子状态被强制改变，应用于{monkey_id}；"
                f"目前状态：坐标{x}, {y}，等级 {target_path1}, {target_path2}, {target_path3}",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "设置速度":
        print(f"执行指令：设置游戏速度")
        change_game_speed(initial_param)
        # 设置1或3倍速
        if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
            write_game_log(
                f"速度调整为{initial_param}倍速",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "强制判定关卡完成":
        print(f"执行指令：强制判定关卡完成")
        config.NOW_GAME_LEVEL += 1
        if config.LOG_FILE_GRANULARITY >= 1:  # 记录日志
            write_game_log(
                f"强制判定关卡完成",
                config.CUSTOM_SAVE_PATH
            )

    elif command == "强制判定本局完成":
        print(f"执行指令：强制判定本局完成")
        config.NOW_GAME_LEVEL = config.MAX_GAME_LEVEL + 1
        if config.LOG_FILE_GRANULARITY >= 1:  # 记录日志
            write_game_log(
                f"强制判定本局完成",
                config.CUSTOM_SAVE_PATH
            )

    else:
        print(f"未知指令或xlsx文件空行: {command}")
        # if config.LOG_FILE_GRANULARITY >= 2:  # 记录日志
        #     write_game_log(
        #         f"读取未知指令或xlsx文件空行",
        #         config.CUSTOM_SAVE_PATH
        #     )


# 从 xlsx文件加载和执行指令
def execute_commands(file_path):
    try:
        # 屏蔽openpyxl格式警告
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            wb = load_workbook(file_path, read_only=True, data_only=True)

            worksheet = wb.worksheets[0] if wb.worksheets else None
            if not worksheet:
                raise ValueError("Excel文件没有有效的工作表")

            # 立即读取所有需要的数据到内存中
            try:
                header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
            except StopIteration:
                raise ValueError("工作表为空")

            field_names = [cell.value.strip() if cell.value else f"未知列_{i}"
                           for i, cell in enumerate(header_row, 1)]

            # 读取所有行数据到列表中
            commands = []
            for row in worksheet.iter_rows(min_row=2):
                row_values = [cell.value if cell.value is not None else "" for cell in row]
                command = {field_names[i]: str(val).strip() for i, val in enumerate(row_values) if i < len(field_names)}
                commands.append(command)

            # 现在可以关闭工作簿了
            # wb.close()

            # 使用已读取的数据继续处理
            level_commands = {}
            current_level = config.NOW_GAME_LEVEL

            for command in commands:
                if command.get("指令") == "等待关卡":
                    current_level = int(command.get("首要参数"))
                if current_level not in level_commands:
                    level_commands[current_level] = []
                level_commands[current_level].append(command)

            min_level = config.START_GAME_LEVEL + 1
            max_level = config.MAX_GAME_LEVEL

            # 自动补全缺失的关卡
            for level in range(min_level, max_level + 2):
                # max_level+2是因为末关也需要重试。
                # 例如点击在100关结束，需要加入等待关卡101才能重试100关。否则100关操作执行后会直接退出本局，无法判断失败
                if level not in level_commands:
                    level_commands[level] = [{
                        "指令": "等待关卡",
                        "首要参数": str(level),
                    }]

            # 按关卡号排序
            level_commands = dict(sorted(level_commands.items()))

            # 调整关前停顿指令位置（必须在对应关卡的第一位）
            for level in level_commands:
                # 分离关前停顿和其他指令
                pause_commands = [cmd for cmd in level_commands[level] if cmd.get("指令") == "关前停顿"]
                other_commands = [cmd for cmd in level_commands[level] if cmd.get("指令") != "关前停顿"]

                # 将关前停顿插入到当前关卡指令的最前面
                level_commands[level] = pause_commands + other_commands

            # 生成关前停顿关卡列表
            levels_with_stops = []
            for level, cmds in level_commands.items():
                if any(cmd.get("指令") == "关前停顿" for cmd in cmds):
                    levels_with_stops.append(level)
            levels_with_stops.sort()  # 确保顺序正确
            config.LEVELS_WITH_STOPS = levels_with_stops

            last_command_list = {}
            # print(level_commands)
            print(config.LEVELS_WITH_STOPS)

            for level, command_list in list(level_commands.items()):
                while config.SCRIPT_STOP == 0:
                    try:
                        print(f"\n正在执行第 {config.NOW_GAME_LEVEL} 关指令...")
                        execute_level_commands(level, command_list)
                        last_command_list = command_list
                        break

                    except LevelRetryException as e:
                        print(f"第 {config.NOW_GAME_LEVEL} 关失败，准备重试本关 ({str(e)})")
                        if config.NOW_DIFFICULTY == "点击":  # （在脚本多处有难度判断，维护时请搜config.NOW_DIFFICULTY）
                            global monkeys, monkeys_temp
                            monkeys = copy.deepcopy(monkeys_temp)
                            retry_this_level(level, last_command_list)
                            print(f"调用重试功能")
                        else:
                            print(f"当前难度为{config.NOW_DIFFICULTY}，无法重试")

            # 现在可以关闭工作簿了
            wb.close()

    except Exception as e:
        print(f"执行命令时出错: {e}")
        if 'wb' in locals():
            wb.close()
        raise


def execute_commands_test_placement(file_path):
    try:
        # 屏蔽openpyxl格式警告
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            wb = load_workbook(file_path, read_only=True, data_only=True)

            worksheet = wb.worksheets[0] if wb.worksheets else None
            if not worksheet:
                raise ValueError("Excel文件没有有效的工作表")

            # 立即读取所有需要的数据到内存中
            try:
                header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
            except StopIteration:
                raise ValueError("工作表为空")

            field_names = [cell.value.strip() if cell.value else f"未知列_{i}"
                           for i, cell in enumerate(header_row, 1)]

            # 读取所有行数据到列表中
            commands = []
            for row in worksheet.iter_rows(min_row=2):
                row_values = [cell.value if cell.value is not None else "" for cell in row]
                command = {field_names[i]: str(val).strip() for i, val in enumerate(row_values) if i < len(field_names)}
                commands.append(command)

            # 使用已读取的数据继续处理
            allowed_commands = {
                '地图难度', '放置猴子', '升级猴子', '急切放置猴子', '急切升级猴子',
                '猴子状态被强制改变', '改变目标', '改变瞄准'
            }

            level_commands = {}
            current_level = config.NOW_GAME_LEVEL

            # 构建关卡指令字典（过滤不需要的指令）
            for command in commands:
                cmd_type = command.get("指令")
                if cmd_type == "等待关卡":
                    current_level = int(command.get("首要参数"))
                    continue  # 跳过该指令不存储

                if cmd_type in allowed_commands:
                    if current_level not in level_commands:
                        level_commands[current_level] = []
                    level_commands[current_level].append(command)

            # 过滤空关卡并按关卡号排序
            level_commands = {k: v for k, v in sorted(level_commands.items()) if len(v) > 0}

            # 根据TEST_LEVEL过滤需要执行的关卡
            filtered_levels = {k: v for k, v in level_commands.items() if k <= config.TEST_LEVEL}

            # 新增：在所有关卡处理完成后替换指令类型
            for level, commands in filtered_levels.items():
                for cmd in commands:
                    if cmd.get("指令") == "急切升级猴子":
                        cmd["指令"] = "升级猴子"
                    elif cmd.get("指令") == "急切放置猴子":
                        cmd["指令"] = "放置猴子"

            last_command_list = {}

            # 现在可以关闭工作簿了
            wb.close()

            # 执行过滤后的关卡指令
            for level, command_list in sorted(filtered_levels.items()):
                while config.SCRIPT_STOP == 0:
                    try:
                        print(f"\n正在执行第 {level} 关指令...")
                        execute_level_commands(level, command_list)
                        last_command_list = command_list
                        break
                    except LevelRetryException as e:
                        print(f"放置测试执行错误({str(e)})")

    except Exception as e:
        print(f"执行命令时出错: {e}")
        if 'wb' in locals():
            wb.close()
        raise


def retry_this_level(level, command_list):
    """
    重试当前关卡的所有指令
    """

    while config.RETRY_TIMES <= config.ALLOWED_RETRY_TIMES and config.NOW_DIFFICULTY == "点击":  # 最多重试次数
        # （在脚本多处有难度判断，维护时请搜config.NOW_DIFFICULTY）
        try:
            print(f"正在重试第 {config.NOW_GAME_LEVEL} 关，当前重试次数：{config.RETRY_TIMES}")
            execute_level_commands(level, command_list)
            return  # 重试成功，退出循环
        except Exception as e:
            config.RETRY_TIMES += 1
            print(f"第 {config.NOW_GAME_LEVEL} 关重试失败 ({str(e)})")
            if config.RETRY_TIMES >= config.ALLOWED_RETRY_TIMES:
                raise GameRestartException(f"第 {config.NOW_GAME_LEVEL} 关重试失败超过限制，准备重新开始游戏！")


def execute_level_commands(level, command_list):
    i = 0  # 使用索引循环，而不是直接使用 `for row in commands`
    while i < len(command_list):
        row = command_list[i]
        # print(f"Command row:\n{row}\nSTOP_HERE:{config.STOP_HERE}")
        if config.SCRIPT_STOP == 1:
            return

        global monkeys, monkeys_temp

        # 如果是等待关卡指令，更新全局关卡上下文
        if row["指令"] == "等待关卡":
            initial_param = row["首要参数"]
            target_level = int(initial_param) if initial_param else 0
            try:
                print(f"\n当前关卡{config.NOW_GAME_LEVEL}操作已完成，等待到第 {target_level} 关...")
                global_text.TEXT_NOW_COMMAND = row
                global_text.renew()
                wait_level(target_level)  # 调用 wait_level 更新 config.NOW_GAME_LEVEL

                # 重要！何时更新猴子状态——关卡已成功完成时，而非该关卡中指令完成时
                monkeys_temp = copy.deepcopy(monkeys)

            except GameRestartException:
                print(f"检测到关卡{level}卡死，准备重新开始游戏！")
                raise  # 抛出异常，通知外层逻辑重新启动

        # 如果是连续的“使用道具”指令，则合并批量处理
        elif row["指令"] == "使用道具":
            initial_param = row["首要参数"]

            if initial_param:
                config.IF_GERALDO_PLACED_LEFT = int(initial_param)

            tool_list = []

            # 合并后续连续的“使用道具”行
            while i < len(command_list) and command_list[i]["指令"] == "使用道具":
                tool_id = command_list[i]["id"]  # 从当前行读取道具 ID
                if tool_id in monkeys:  # 确保 ID 存在于猴子字典中
                    tool_to_use = monkeys[tool_id]
                    tool_list.append([
                        tool_to_use.x,  # 从字典中提取 x 坐标
                        tool_to_use.y,  # 从字典中提取 y 坐标
                        tool_to_use.key.lower(),  # 从字典中提取道具名称并转小写
                        tool_to_use.monkey_id,  # 提取ID方便日志记录
                    ])
                else:
                    print(f"警告：道具 ID {tool_id} 未在初始状态中找到，跳过该指令。")
                i += 1  # 移动到下一行

            # 调用批量使用道具函数
            if tool_list:
                print(f"执行指令：使用杰拉尔多道具 (关卡: {config.NOW_GAME_LEVEL})")
                global_text.TEXT_NOW_COMMAND = row
                global_text.renew()
                # print(tool_list)
                use_geraldo_tool(tool_list)

            continue  # 跳过递增 i 的操作，因为已经在上面处理了

        else:
            # 执行其他非“使用道具”的指令
            if row['指令']:
                global_text.TEXT_NOW_COMMAND = row
                global_text.renew()
            execute_command(row)
        i += 1  # 处理下一行指令

    print("\n所有指令执行完成！")


# 主游戏逻辑
def play_game(file_path, highly_custom=0):
    """
    主游戏逻辑，处理游戏重启和指令执行
    """
    if highly_custom:  # 高度自定义模式，不尝试重新开始游戏
        config.IF_GERALDO_PLACED_LEFT = config.IF_SHOP_OPENED = 0
        load_initial_state(file_path)
        execute_commands(file_path)  # 从 xlsx 文件加载并执行指令
        print("游戏完成！")

    else:  # 一般模式，会尝试重新开始游戏
        while True:  # 外层循环用于重启整个游戏
            try:
                print("开始新一局游戏...")
                config.IF_GERALDO_PLACED_LEFT = config.IF_SHOP_OPENED = 0
                load_initial_state(file_path)
                execute_commands(file_path)  # 从 xlsx 文件加载并执行指令
                print("游戏完成！")
                break  # 如果游戏顺利完成，退出循环

            except GameRestartException:
                print("重新开始游戏！")
                continue  # 捕获自定义异常，重新开始游戏逻辑


def play_game_test_placement(file_path):
    """
    主游戏逻辑，处理游戏重启和指令执行
    """
    print("开始测试摆放...")
    config.IF_GERALDO_PLACED_LEFT = config.IF_SHOP_OPENED = 0
    load_initial_state(file_path)
    execute_commands_test_placement(file_path)  # 从 xlsx 文件加载并执行指令
    print("摆放完成！")

# play_game('e')
